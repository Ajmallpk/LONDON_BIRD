from django.shortcuts import render
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from django.http import FileResponse
from datetime import datetime, timedelta
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter  
from checkout.models import Order
from decimal import Decimal
from django.db.models import Sum
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether

@staff_member_required
def sales_report(request):
    # Set end_date to the current time (timezone-aware)
    end_date = timezone.now()
    # Default to last 30 days
    start_date = end_date - timedelta(days=30)

    date_filter = request.GET.get('date_filter', 'custom')
    custom_start = request.GET.get('start_date')
    custom_end = request.GET.get('end_date')

    if date_filter == 'daily':
        start_date = end_date - timedelta(days=1)
    elif date_filter == 'weekly':
        start_date = end_date - timedelta(days=7)
    elif date_filter == 'monthly':
        start_date = end_date - timedelta(days=30)
    elif date_filter == 'yearly':
        start_date = end_date - timedelta(days=365)
    elif date_filter == 'custom' and custom_start and custom_end:
        try:
            # Parse the dates and make them timezone-aware
            start_date = datetime.strptime(custom_start, '%Y-%m-%d')
            end_date = datetime.strptime(custom_end, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date)
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
        except ValueError:
            pass

    # Fetch orders and sort by created_at in descending order (newest first)
    orders = Order.objects.filter(
        created_at__range=[start_date, end_date],
    ).order_by('-created_at')

    # Debugging: Print the orders to see what’s being fetched
    print(f"Start Date: {start_date}, End Date: {end_date}")
    print(f"Orders found: {orders.count()}")
    for order in orders:
        print(f"Order {order.id}: Created At: {order.created_at}, Status: {order.status}, Total Price: {order.total_price}")

    sales_count = orders.count()
    total_price = orders.aggregate(Sum('total_price'))['total_price__sum'] or Decimal('0.00')
    orders_with_discount = 0
    total_product_discount = Decimal('0.00')
    total_coupon_discount = orders.aggregate(Sum('discount_coupon_amount'))['discount_coupon_amount__sum'] or Decimal('0.00')

    for order in orders:
        # Calculate product discount using final_offer_price if available
        product_discount = order.product_discount_amount
        if product_discount == Decimal('0.00'):
            # Fallback: Calculate discount based on final_offer_price vs original price
            product_discount = Decimal('0.00')
            for item in order.items.all():
                original_price = item.get_original_unit_price() * item.quantity
                # Use final_offer_price if set; otherwise, use item.price
                final_price = (item.final_offer_price * item.quantity) if item.final_offer_price > 0 else (item.price * item.quantity)
                discount = original_price - final_price
                product_discount += max(discount, Decimal('0.00'))  # Ensure no negative discounts
        
        total_product_discount += product_discount

        has_discount = product_discount > 0 or (order.discount_coupon_amount or 0) > 0
        if has_discount:
            orders_with_discount += 1

        # Debugging: Print discount values for each order
        print(f"Order {order.id}: Product Discount: {product_discount}, Coupon Discount: {order.discount_coupon_amount}")
        for item in order.items.all():
            print(f"  Item {item.id}: Original Price (Unit): {item.get_original_unit_price()}, Final Offer Price: {item.final_offer_price}, Item Price: {item.price}, Quantity: {item.quantity}")

    total_discount_combined = (total_product_discount or 0) + (total_coupon_discount or 0)
    net_price = total_price - total_discount_combined if total_price else Decimal('0.00')

    orders_data = []
    for order in orders:
        product_discount = order.product_discount_amount
        if product_discount == Decimal('0.00'):
            # Same calculation as above for individual order data
            product_discount = Decimal('0.00')
            for item in order.items.all():
                original_price = item.get_original_unit_price() * item.quantity
                final_price = (item.final_offer_price * item.quantity) if item.final_offer_price > 0 else (item.price * item.quantity)
                discount = original_price - final_price
                product_discount += max(discount, Decimal('0.00'))
        
        net = order.total_price - product_discount - (order.discount_coupon_amount or 0)
        orders_data.append({
            'order': order,
            'product_discount': product_discount,
            'net_amount': net,
            'has_discount': product_discount > 0 or (order.discount_coupon_amount or 0) > 0,
        })

    download_format = request.GET.get('download', '')
    if download_format:
        if download_format == 'pdf':
            return generate_pdf_report(orders_data, start_date, end_date, sales_count, total_price, orders_with_discount, total_product_discount, total_coupon_discount, total_discount_combined, net_price)
        elif download_format == 'excel':
            return generate_excel_report(orders_data, start_date, end_date, sales_count, total_price, orders_with_discount, total_product_discount, total_coupon_discount, total_discount_combined, net_price)

    context = {
        'orders_data': orders_data,
        'start_date': start_date.date(),
        'end_date': end_date.date(),
        'sales_count': sales_count,
        'total_price': total_price,
        'orders_with_discount': orders_with_discount,
        'total_product_discount': total_product_discount,
        'total_coupon_discount': total_coupon_discount,
        'total_discount_combined': total_discount_combined,
        'net_price': net_price,
        'date_filter': date_filter,
    }
    return render(request, 'sales_report.html', context)

def generate_pdf_report(orders_data, start_date, end_date, sales_count, total_price, orders_with_discount, total_product_discount, total_coupon_discount, total_discount_combined, net_price):
    buffer = io.BytesIO()
    # Set page margins and size
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.5 * inch,
        rightMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch
    )
    elements = []

    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name='Title',
        fontSize=16,
        leading=20,
        alignment=1,  # Center
        spaceAfter=20,
        fontName='Helvetica-Bold'
    )
    cell_style = ParagraphStyle(
        name='CellStyle',
        fontSize=8,
        leading=10,
        alignment=0,  # Left
        fontName='Helvetica',
        wordWrap='CJK'  # Enable text wrapping
    )
    cell_style_right = ParagraphStyle(
        name='CellStyleRight',
        fontSize=8,
        leading=10,
        alignment=2,  # Right
        fontName='Helvetica',
        wordWrap='CJK'
    )

    # Title
    elements.append(Paragraph(f"Sales Report ({start_date.date()} to {end_date.date()})", title_style))
    elements.append(Spacer(1, 0.25 * inch))

    # Summary Table
    summary_col_widths = [4.5 * inch, 2.5 * inch]  # Total 7in (leaving some buffer)
    summary_data = [
        ['Summary', ''],
        ['Total Sales Count', str(sales_count)],
        ['Total Order Amount (₹)', f"{total_price:.2f}"],
        ['Orders with Discount Applied', str(orders_with_discount)],
        ['Total Product/Category Discount (₹)', f"{total_product_discount:.2f}"],
        ['Total Coupon Discount (₹)', f"{total_coupon_discount:.2f}"],
        ['Total Discount (₹)', f"{total_discount_combined:.2f}"],
        ['Net Amount (₹)', f"{net_price:.2f}"],
    ]
    # Wrap summary table content in Paragraphs for better control
    summary_data_wrapped = [
        [Paragraph(cell if isinstance(cell, str) else str(cell), cell_style if col == 0 else cell_style_right) for col, cell in enumerate(row)]
        for row in summary_data
    ]
    summary_table = Table(summary_data_wrapped, colWidths=summary_col_widths, rowHeights=[None] * len(summary_data))
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.5 * inch))

    # Orders Table
    # Total width = 7in (leaving some buffer)
    col_widths = [
        1.2 * inch,  # Order ID
        1.3 * inch,  # Date
        0.9 * inch,  # Total Amount
        0.9 * inch,  # Discount Applied
        0.9 * inch,  # Product Discount
        0.9 * inch,  # Coupon Discount
        0.9 * inch   # Net Amount
    ]
    # Debug: Print column widths to verify total
    total_width = sum(col_widths)
    print(f"Total Orders Table Width: {total_width / inch} inches (should be <= 7.5 inches)")

    data = [['Order ID', 'Date', 'Total Amount (₹)', 'Discount Applied', 'Product Discount (₹)', 'Coupon Discount (₹)', 'Net Amount (₹)']]
    for entry in orders_data:
        order = entry['order']
        row = [
            str(order.id),
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            f"{order.total_price:.2f}",
            'Yes' if entry['has_discount'] else 'No',
            f"{entry['product_discount']:.2f}",
            f"{order.discount_coupon_amount or 0:.2f}",
            f"{entry['net_amount']:.2f}",
        ]
        # Debug: Print content lengths to identify potential overflow
        print(f"Order {order.id}: {[(col, len(str(val))) for col, val in enumerate(row)]}")
        data.append(row)

    # Wrap content in Paragraphs to enable text wrapping
    data_wrapped = []
    for row in data:
        wrapped_row = []
        for col, cell in enumerate(row):
            # Use right-aligned style for numerical columns (2 onwards, except col 3)
            style = cell_style_right if col >= 2 and col != 3 else cell_style
            wrapped_row.append(Paragraph(str(cell), style))
        data_wrapped.append(wrapped_row)

    # Split data into chunks to handle page breaks
    chunk_size = 20  # Number of rows per page
    for i in range(0, len(data_wrapped), chunk_size):
        chunk = data_wrapped[i:i + chunk_size]
        # Set row heights to accommodate wrapped text
        row_heights = [0.3 * inch if r == 0 else None for r in range(len(chunk))]
        table = Table(chunk, colWidths=col_widths, rowHeights=row_heights, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        elements.append(KeepTogether(table))
        if i + chunk_size < len(data_wrapped):
            elements.append(Spacer(1, 0.25 * inch))

    doc.build(elements)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='sales_report.pdf')

def generate_excel_report(orders_data, start_date, end_date, sales_count, total_price, orders_with_discount, total_product_discount, total_coupon_discount, total_discount_combined, net_price):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    ws.append([f"Sales Report ({start_date.date()} to {end_date.date()})"])
    ws.merge_cells('A1:G1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.append(['Summary'])
    ws.append(['Total Sales Count', sales_count])
    ws.append(['Total Order Amount (₹)', f"{total_price:.2f}"])
    ws.append(['Orders with Discount Applied', orders_with_discount])
    ws.append(['Total Product/Category Discount (₹)', f"{total_product_discount:.2f}"])
    ws.append(['Total Coupon Discount (₹)', f"{total_coupon_discount:.2f}"])
    ws.append(['Total Discount (₹)', f"{total_discount_combined:.2f}"])
    ws.append(['Net Amount (₹)', f"{net_price:.2f}"])
    ws.append([])

    headers = ['Order ID', 'Date', 'Total Amount (₹)', 'Discount Applied', 'Product Discount (₹)', 'Coupon Discount (₹)', 'Net Amount (₹)']
    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    for entry in orders_data:
        order = entry['order']
        ws.append([
            order.id,
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            f"{order.total_price:.2f}",
            'Yes' if entry['has_discount'] else 'No',
            f"{entry['product_discount']:.2f}",
            f"{order.discount_coupon_amount or 0:.2f}",
            f"{entry['net_amount']:.2f}",
        ])

    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='sales_report.xlsx')